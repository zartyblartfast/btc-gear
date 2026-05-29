#!/usr/bin/env python3
"""Build the BTC Leveraged Accumulation spreadsheet as .xlsx"""

import sys
sys.path.insert(0, r'C:\Users\clive\miniconda3\Lib\site-packages')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
import math

wb = Workbook()

# ============================================================
# STYLES
# ============================================================
hdr_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
hdr_fill = PatternFill(start_color='023D4B', end_color='023D4B', fill_type='solid')
sec_font = Font(name='Calibri', size=12, bold=True, color='023D4B')
sec_fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')
lbl_font = Font(name='Calibri', size=11, color='333333')
val_font = Font(name='Calibri', size=11, bold=True)
desc_font = Font(name='Calibri', size=10, color='888888')
input_bg = PatternFill(start_color='F6F6F6', end_color='F6F6F6', fill_type='solid')
title_font = Font(name='Calibri', size=18, bold=True, color='023D4B')
thin = Border(left=Side('thin','00DDDDDD'), right=Side('thin','00DDDDDD'),
              top=Side('thin','00DDDDDD'), bottom=Side('thin','00DDDDDD'))
green_bg = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
amber_bg = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
red_bg = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')

def hdr(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = thin
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def cell(ws, r, c, v=None, fmt=None, font=None, fill=None, align='center'):
    cl = ws.cell(row=r, column=c, value=v)
    cl.border = thin; cl.alignment = Alignment(horizontal=align, vertical='center')
    if fmt: cl.number_format = fmt
    if font: cl.font = font
    if fill: cl.fill = fill
    return cl

# ============================================================
# TAB 1: INPUTS
# ============================================================
ws = wb.active
ws.title = "Inputs"
ws.sheet_properties.tabColor = "023D4B"
ws.column_dimensions['A'].width = 36
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 13
ws.column_dimensions['D'].width = 55

ws.merge_cells('A1:D1')
cell(ws, 1, 1, 'BTC Leveraged Accumulation & Income Model', font=title_font, align='left')

def section(ws, r, title):
    ws.merge_cells(f'A{r}:D{r}')
    cell(ws, r, 1, title, font=sec_font, fill=sec_fill, align='left')
    return r+1

def inp(ws, r, name, val, unit, desc):
    cell(ws, r, 1, name, font=lbl_font, fill=input_bg, align='left')
    cell(ws, r, 2, val, font=val_font)
    cell(ws, r, 3, unit, font=desc_font)
    cell(ws, r, 4, desc, font=desc_font, align='left')
    return r+1

rr = 3
rr = section(ws, rr, 'BTC POSITION')
rr = inp(ws, rr, 'BTC Holdings', 2.0, 'BTC', 'Total Bitcoin used as collateral')
rr = inp(ws, rr, 'Current BTC Price', 75000, 'USD', 'Spot price at model start')
rr += 1

rr = section(ws, rr, 'BORROWING PARAMETERS')
rr = inp(ws, rr, 'LTV Target', 0.35, '%', 'Target LTV. Ledn max: 50%.')
rr = inp(ws, rr, 'Borrow APR', 0.1149, '%', 'Annual interest. Ledn: 9.75–11.49%')
rr = inp(ws, rr, 'Origination Fee', 0.0, '%', 'Ledn: 0%')
rr = inp(ws, rr, 'Annual Platform Fee', 0.0, '%', 'Ledn: 0%')
rr = inp(ws, rr, 'Margin Call LTV', 0.70, '%', 'Ledn: 70% LTV')
rr = inp(ws, rr, 'Liquidation LTV', 0.80, '%', 'Ledn: 80% LTV')
rr = inp(ws, rr, 'Safety Margin', 0.10, 'pp', 'Warning = margin_call − safety')
rr += 1

rr = section(ws, rr, 'MODE')
rr = inp(ws, rr, 'Mode', 'Accumulation', 'toggle', '"Accumulation" or "Income"')
rr = inp(ws, rr, 'Withdrawal Rule', '% of equity gain', 'toggle', '"Fixed $" or "% of equity gain"')
rr = inp(ws, rr, 'Withdrawal Amount', 0.50, 'USD or %', '50% of equity gain or $X/year')
rr += 1

rr = section(ws, rr, 'REBALANCING')
rr = inp(ws, rr, 'Rebalance Rule', 'Maintain LTV', 'choice', '"Maintain LTV", "Never Increase", "Dynamic"')
rr += 1

rr = section(ws, rr, 'PRICE SCENARIO ANCHORS')
rr = inp(ws, rr, 'Pessimistic 2030 Anchor', 100000, 'USD', 'BTC price in 2030 — pessimistic')
rr = inp(ws, rr, 'Median 2030 Anchor', 500000, 'USD', 'BTC price in 2030 — median')
rr = inp(ws, rr, 'Optimistic 2030 Anchor', 1000000, 'USD', 'BTC price in 2030 — optimistic')
rr = inp(ws, rr, 'Post-2030 Growth Decay', 0.30, 'frac', 'Phase 2 CAGR as fraction of Phase 1')
rr = inp(ws, rr, 'Cycle Amplitude', 0.40, 'frac', 'Peak-to-trough amplitude')
rr = inp(ws, rr, 'Amplitude Decay Per Cycle', 0.15, 'frac', 'Reduction per 4-year cycle')
rr += 1

rr = section(ws, rr, 'MODEL')
rr = inp(ws, rr, 'Start Year', 2025, 'year', 'First projection year')
rr = inp(ws, rr, 'Projection Length', 20, 'years', 'Years to model')
rr += 1

rr = section(ws, rr, 'INFLATION')
rr = inp(ws, rr, 'Show Real Values', 'Yes', 'toggle', 'Display inflation-adjusted values')
rr = inp(ws, rr, 'Annual Inflation Rate', 0.03, '%', 'Assumed USD inflation rate')

# Data validation dropdowns for cells with predefined options
dv_mode = DataValidation(type="list", formula1='"Accumulation,Income"', allow_blank=True)
dv_mode.error = "Must be 'Accumulation' or 'Income'"
dv_mode.errorTitle = "Invalid Mode"
ws.add_data_validation(dv_mode)
dv_mode.add('B17')

dv_withdraw = DataValidation(type="list", formula1='"Fixed $,% of equity gain"', allow_blank=True)
dv_withdraw.error = "Must be 'Fixed $' or '% of equity gain'"
dv_withdraw.errorTitle = "Invalid Withdrawal Rule"
ws.add_data_validation(dv_withdraw)
dv_withdraw.add('B18')

dv_rebalance = DataValidation(type="list", formula1='"Maintain LTV,Never Increase,Dynamic"', allow_blank=True)
dv_rebalance.error = "Must be 'Maintain LTV', 'Never Increase', or 'Dynamic'"
dv_rebalance.errorTitle = "Invalid Rebalance Rule"
ws.add_data_validation(dv_rebalance)
dv_rebalance.add('B22')

dv_real = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
dv_real.error = "Must be 'Yes' or 'No'"
dv_real.errorTitle = "Invalid Option"
ws.add_data_validation(dv_real)
dv_real.add('B37')

# Store cell references for formula construction
# These are the EXACT cells users enter data into — MUST include sheet name
# because formulas live on other tabs
BTC_HOLDINGS = 'Inputs!$B$4'
BTC_PRICE    = 'Inputs!$B$5'
LTV          = 'Inputs!$B$8'
BORROW_APR   = 'Inputs!$B$9'
ORIG_FEE     = 'Inputs!$B$10'
PLAT_FEE     = 'Inputs!$B$11'
MARGIN_CALL  = 'Inputs!$B$12'
LIQ_LTV      = 'Inputs!$B$13'
SAFETY_MARGIN = 'Inputs!$B$14'
MODE         = 'Inputs!$B$17'
WITHDRAW_RULE = 'Inputs!$B$18'
WITHDRAW_AMT = 'Inputs!$B$19'
REBAL_RULE   = 'Inputs!$B$22'
ANCHOR_PESS  = 'Inputs!$B$25'
ANCHOR_MED   = 'Inputs!$B$26'
ANCHOR_OPT   = 'Inputs!$B$27'
GROWTH_DECAY = 'Inputs!$B$28'
CYCLE_AMP    = 'Inputs!$B$29'
AMP_DECAY    = 'Inputs!$B$30'
START_YEAR   = 'Inputs!$B$33'
PROJ_LENGTH  = 'Inputs!$B$34'
SHOW_REAL    = 'Inputs!$B$37'
INFLATION    = 'Inputs!$B$38'

print("Tab 1 (Inputs) created.")

# ============================================================
# TAB 2: PRICE PROJECTION
# ============================================================
ws2 = wb.create_sheet("Price Projection")
ws2.sheet_properties.tabColor = "1A6B7A"
ws2.column_dimensions['A'].width = 6
for c in range(2, 14):
    ws2.column_dimensions[get_column_letter(c)].width = 18

# Headers
headers = ['t', 'Year', 'Pess Trend', 'Pess Cycle Mult', 'Pess Price',
           'Med Trend', 'Med Cycle Mult', 'Med Price',
           'Opt Trend', 'Opt Cycle Mult', 'Opt Price', 'Med YoY %']
for i, h in enumerate(headers, 1):
    cell(ws2, 1, i, h, font=hdr_font, fill=hdr_fill)
hdr(ws2, 1, len(headers))

N = 21  # years 0-20 (2025-2045)
CHART_ROWS = N + 1  # header + data rows for chart ranges

for t in range(N):
    r = t + 2  # data starts row 2
    yr = 2025 + t

    cell(ws2, r, 1, t)
    cell(ws2, r, 2, yr)

    # ---- Pessimistic ----
    # CAGR1 = (Anchor/CurrentPrice)^(1/5) - 1  (expression only, no leading =)
    cagr1_pess = f'({ANCHOR_PESS}/{BTC_PRICE})^(1/5)-1'
    # CAGR2 = CAGR1 * Growth_Decay
    cagr2_pess = f'({cagr1_pess})*{GROWTH_DECAY}'
    # Trend: if t<=5: Price*(1+CAGR1)^t else: Anchor*(1+CAGR2)^(t-5)
    trend_pess = f'=IF(A{r}<=5,{BTC_PRICE}*(1+({cagr1_pess}))^A{r},{ANCHOR_PESS}*(1+({cagr2_pess}))^(A{r}-5))'
    cell(ws2, r, 3, trend_pess, fmt='$#,##0')

    # Normalized cycle: Raw = 1 + amp*cos; divide by (1+amp) so t=0 always = 1.0
    # effective_amp = base_amp * (1-amp_decay)^floor(t/4)
    cycle_pess = f'=(1+({CYCLE_AMP}*(1-{AMP_DECAY})^FLOOR(A{r}/4,1))*COS(2*PI()*A{r}/4))/(1+{CYCLE_AMP}*(1-{AMP_DECAY})^FLOOR(A{r}/4,1))'
    cell(ws2, r, 4, cycle_pess, fmt='0.0000')

    # Price = Trend * Cycle_Multiplier
    price_pess = f'=C{r}*D{r}'
    cell(ws2, r, 5, price_pess, fmt='$#,##0')

    # ---- Median ----
    cagr1_med = f'({ANCHOR_MED}/{BTC_PRICE})^(1/5)-1'
    cagr2_med = f'({cagr1_med})*{GROWTH_DECAY}'
    trend_med = f'=IF(A{r}<=5,{BTC_PRICE}*(1+({cagr1_med}))^A{r},{ANCHOR_MED}*(1+({cagr2_med}))^(A{r}-5))'
    cell(ws2, r, 6, trend_med, fmt='$#,##0')

    cycle_med = f'=(1+({CYCLE_AMP}*(1-{AMP_DECAY})^FLOOR(A{r}/4,1))*COS(2*PI()*A{r}/4))/(1+{CYCLE_AMP}*(1-{AMP_DECAY})^FLOOR(A{r}/4,1))'
    cell(ws2, r, 7, cycle_med, fmt='0.0000')

    price_med = f'=F{r}*G{r}'
    cell(ws2, r, 8, price_med, fmt='$#,##0')

    # ---- Optimistic ----
    cagr1_opt = f'({ANCHOR_OPT}/{BTC_PRICE})^(1/5)-1'
    cagr2_opt = f'({cagr1_opt})*{GROWTH_DECAY}'
    trend_opt = f'=IF(A{r}<=5,{BTC_PRICE}*(1+({cagr1_opt}))^A{r},{ANCHOR_OPT}*(1+({cagr2_opt}))^(A{r}-5))'
    cell(ws2, r, 9, trend_opt, fmt='$#,##0')

    cycle_opt = f'=(1+({CYCLE_AMP}*(1-{AMP_DECAY})^FLOOR(A{r}/4,1))*COS(2*PI()*A{r}/4))/(1+{CYCLE_AMP}*(1-{AMP_DECAY})^FLOOR(A{r}/4,1))'
    cell(ws2, r, 10, cycle_opt, fmt='0.0000')

    price_opt = f'=I{r}*J{r}'
    cell(ws2, r, 11, price_opt, fmt='$#,##0')

    # YoY change (median)
    if t == 0:
        cell(ws2, r, 12, '', fmt='0.0%')
    else:
        cell(ws2, r, 12, f'=IF(H{r-1}=0,"",H{r}/H{r-1}-1)', fmt='0.0%')

print("Tab 2 (Price Projection) created.")

# ============================================================
# TAB 3: LEVERAGED POSITION
# ============================================================
ws3 = wb.create_sheet("Leveraged Position")
ws3.sheet_properties.tabColor = "1A6B7A"
cols3 = [
    'Year', 'BTC Price', 'Collateral BTC', 'Collateral Value', 'Target LTV',
    'Debt Start', 'Pre-Interest LTV', 'Annual Interest', 'Debt After Interest',
    'Target Debt', 'Required Repayment', 'Rebalance Borrowing', 'Rebalance Repayment',
    'Repayment Capped?', 'External Top-Up Required', 'Debt End',
    'BTC Bought (Rebal)', 'BTC Sold (Repayment)', 'Total Lev BTC Before Income',
    'Gross Value Before Income', 'Net Equity Before Income', 'Equity Gain USD',
    'Requested Income USD', 'Income Withdrawn USD', 'BTC Sold for Income',
    'Total Lev BTC End', 'Gross Position Value End', 'Net Equity USD End',
    'Net Equity BTC End', 'Annual Fee', 'Total Annual Cost', 'Price Change $',
    'Mark-to-Market Return', 'Effective LTV End', 'Margin Call Thr',
    'Liquidation Thr', 'Safety Buffer', 'Risk Status', 'Renewal Risk',
    'LTV Suggestion', 'Purchased BTC Before Repay', 'Purchased BTC Before Income',
    'Purchased BTC End'
]
for i, h in enumerate(cols3, 1):
    cell(ws3, 1, i, h, font=hdr_font, fill=hdr_fill)
    ws3.column_dimensions[get_column_letter(i)].width = 16
hdr(ws3, 1, len(cols3))

for t in range(N):
    r = t + 2
    yr = 2025 + t
    cell(ws3, r, 1, yr)

    # BTC price = from Tab2 col H (median price)
    cell(ws3, r, 2, f"='Price Projection'!H{r}", fmt='$#,##0')

    # Collateral BTC = original holdings. Constant in v1 unless liquidation occurs.
    cell(ws3, r, 3, f'={BTC_HOLDINGS}', fmt='0.00000000')

    # Collateral Value and Target LTV
    cell(ws3, r, 4, f'=B{r}*C{r}', fmt='$#,##0')
    cell(ws3, r, 5, f'={LTV}', fmt='0.0%')

    # Debt lifecycle in canonical order
    if t == 0:
        cell(ws3, r, 6, f'=D{r}*E{r}', fmt='$#,##0')  # Debt start/setup debt
        cell(ws3, r, 8, 0, fmt='$#,##0')              # Annual interest
    else:
        cell(ws3, r, 6, f'=P{r-1}', fmt='$#,##0')
        cell(ws3, r, 8, f'=F{r}*{BORROW_APR}', fmt='$#,##0')
    cell(ws3, r, 7, f'=IF(D{r}>0,F{r}/D{r},0)', fmt='0.0%')
    cell(ws3, r, 9, f'=F{r}+H{r}', fmt='$#,##0')

    # Price change before target-debt formula because Dynamic uses YoY move
    if t == 0:
        cell(ws3, r, 32, 0, fmt='$#,##0')
        yoy_expr = '0'
    else:
        cell(ws3, r, 32, f'=B{r}-B{r-1}', fmt='$#,##0')
        yoy_expr = f'IF(B{r-1}=0,0,B{r}/B{r-1}-1)'

    # Purchased BTC available before repayment excludes original collateral.
    if t == 0:
        initial_borrowed_btc = f'({BTC_HOLDINGS}*{LTV}*{BTC_PRICE}/B{r})'
        cell(ws3, r, 41, f'={initial_borrowed_btc}', fmt='0.00000000')
    else:
        cell(ws3, r, 41, f'=AQ{r-1}', fmt='0.00000000')

    # Target debt by selected rebalance rule.
    maintain_target = f'D{r}*E{r}'
    never_increase = f'IF(G{r}<={MARGIN_CALL}-{SAFETY_MARGIN},I{r},{maintain_target})'
    dynamic_target = f'IF(({yoy_expr})>-0.10,I{r},IF(({yoy_expr})>=-0.30,{maintain_target},IF(({yoy_expr})>=-0.50,D{r}*MAX(G{r}-0.10,0.10),0)))'
    target_debt = f'=IF({REBAL_RULE}="Never Increase",{never_increase},IF({REBAL_RULE}="Dynamic",{dynamic_target},{maintain_target}))'
    cell(ws3, r, 10, target_debt, fmt='$#,##0')

    # Required repayment and rebalance actions. Repayment is capped to purchased BTC available.
    cell(ws3, r, 11, f'=MAX(0,I{r}-J{r})', fmt='$#,##0')
    if t == 0:
        cell(ws3, r, 12, 0, fmt='$#,##0')
    else:
        cell(ws3, r, 12, f'=IF({REBAL_RULE}="Never Increase",0,MAX(0,J{r}-I{r}))', fmt='$#,##0')
    cell(ws3, r, 13, f'=MIN(K{r},AO{r}*B{r})', fmt='$#,##0')
    cell(ws3, r, 14, f'=K{r}>M{r}')
    cell(ws3, r, 15, f'=MAX(0,K{r}-M{r})', fmt='$#,##0')
    cell(ws3, r, 16, f'=I{r}+L{r}-M{r}', fmt='$#,##0')

    # BTC position before income
    cell(ws3, r, 17, f'=IF(B{r}>0,L{r}/B{r},0)', fmt='0.00000000')
    cell(ws3, r, 18, f'=IF(B{r}>0,M{r}/B{r},0)', fmt='0.00000000')
    if t == 0:
        cell(ws3, r, 19, f'=C{r}+AO{r}+Q{r}-R{r}', fmt='0.00000000')
    else:
        cell(ws3, r, 19, f'=Z{r-1}+Q{r}-R{r}', fmt='0.00000000')
    cell(ws3, r, 20, f'=S{r}*B{r}', fmt='$#,##0')
    cell(ws3, r, 21, f'=T{r}-P{r}', fmt='$#,##0')
    if t == 0:
        cell(ws3, r, 22, 0, fmt='$#,##0')
    else:
        cell(ws3, r, 22, f'=U{r}-AB{r-1}', fmt='$#,##0')

    # Income request and cap. Income is based on pre-income equity gain, then capped to purchased BTC.
    cell(ws3, r, 23,
         f'=IF(OR({MODE}="Accumulation",V{r}<=0),0,IF({WITHDRAW_RULE}="% of equity gain",V{r}*{WITHDRAW_AMT},{WITHDRAW_AMT}))',
         fmt='$#,##0')

    # Purchased BTC after repayment / before income.
    cell(ws3, r, 42, f'=AO{r}+Q{r}-R{r}', fmt='0.00000000')

    cell(ws3, r, 25, f'=IF({MODE}="Income",MIN(IF(B{r}>0,W{r}/B{r},0),AP{r}),0)', fmt='0.00000000')
    cell(ws3, r, 24, f'=Y{r}*B{r}', fmt='$#,##0')
    cell(ws3, r, 26, f'=S{r}-Y{r}', fmt='0.00000000')
    cell(ws3, r, 43, f'=AP{r}-Y{r}', fmt='0.00000000')

    # Post-income equity fields
    cell(ws3, r, 27, f'=Z{r}*B{r}', fmt='$#,##0')
    cell(ws3, r, 28, f'=AA{r}-P{r}', fmt='$#,##0')
    cell(ws3, r, 29, f'=IF(B{r}>0,AB{r}/B{r},0)', fmt='0.00000000')

    # Costs and diagnostics
    cell(ws3, r, 30, f'=D{r}*{PLAT_FEE}', fmt='$#,##0')
    orig = f'+P{r}*{ORIG_FEE}' if t == 0 else ''
    cell(ws3, r, 31, f'=H{r}+AD{r}{orig}', fmt='$#,##0')
    if t == 0:
        cell(ws3, r, 33, 0, fmt='$#,##0')
    else:
        cell(ws3, r, 33, f'=Z{r-1}*AF{r}-AE{r}', fmt='$#,##0')

    # Risk fields. Liquidation/margin checks use pre-interest LTV for renewal risk;
    # end-of-year risk status displays the final effective LTV.
    cell(ws3, r, 34, f'=IF(D{r}>0,P{r}/D{r},0)', fmt='0.0%')
    cell(ws3, r, 35, f'={MARGIN_CALL}', fmt='0.0%')
    cell(ws3, r, 36, f'={LIQ_LTV}', fmt='0.0%')
    cell(ws3, r, 37, f'=IF(D{r}>0,AJ{r}-AH{r},0)', fmt='0.00%')
    cell(ws3, r, 38, f'=IF(G{r}>={LIQ_LTV},"LIQUIDATED",IF(AH{r}>={MARGIN_CALL},"MARGIN CALL",IF(AH{r}>={MARGIN_CALL}-{SAFETY_MARGIN},"WARNING","SAFE")))')
    cell(ws3, r, 39, f'=G{r}>={MARGIN_CALL}')
    cell(ws3, r, 40, f'=IF(AH{r}=0,E{r},J{r}/D{r})', fmt='0.0%')

print("Tab 3 (Leveraged Position) created.")

# ============================================================
# TAB 4: INCOME / ACCUMULATION
# ============================================================
ws4 = wb.create_sheet("Income-Accumulation")
ws4.sheet_properties.tabColor = "1A6B7A"
cols4 = [
    'Year', 'Equity Gain USD', 'Requested Income USD', 'Income Withdrawn USD',
    'BTC Sold for Income', 'Income Capped?', 'Cumulative Income',
    'Cumulative BTC Sold Income', 'BTC from Rebalancing', 'Total Lev BTC End',
    'Net BTC After Debt', 'BTC Accum Mult', 'Passive Hold BTC',
    'Cumul Outperformance', 'Annual Inc % of Capital', 'Purchased BTC Available',
    'Sustainable?'
]
for i, h in enumerate(cols4, 1):
    cell(ws4, 1, i, h, font=hdr_font, fill=hdr_fill)
    ws4.column_dimensions[get_column_letter(i)].width = 16
hdr(ws4, 1, len(cols4))

for t in range(N):
    r = t + 2
    yr = 2025 + t
    cell(ws4, r, 1, yr)
    cell(ws4, r, 2, f"='Leveraged Position'!V{r}", fmt='$#,##0')
    cell(ws4, r, 3, f"='Leveraged Position'!W{r}", fmt='$#,##0')
    cell(ws4, r, 4, f"='Leveraged Position'!X{r}", fmt='$#,##0')
    cell(ws4, r, 5, f"='Leveraged Position'!Y{r}", fmt='0.00000000')
    cell(ws4, r, 6, f'=D{r}<C{r}')
    if t == 0:
        cell(ws4, r, 7, f'=D{r}', fmt='$#,##0')
        cell(ws4, r, 8, f'=E{r}', fmt='0.00000000')
    else:
        cell(ws4, r, 7, f'=G{r-1}+D{r}', fmt='$#,##0')
        cell(ws4, r, 8, f'=H{r-1}+E{r}', fmt='0.00000000')
    cell(ws4, r, 9, f"='Leveraged Position'!Q{r}", fmt='0.00000000')
    cell(ws4, r, 10, f"='Leveraged Position'!Z{r}", fmt='0.00000000')
    cell(ws4, r, 11, f"='Leveraged Position'!AC{r}", fmt='0.00000000')
    cell(ws4, r, 12, f'=IF({BTC_HOLDINGS}>0,K{r}/{BTC_HOLDINGS},0)', fmt='0.00x')
    cell(ws4, r, 13, f'={BTC_HOLDINGS}', fmt='0.00000000')
    cell(ws4, r, 14, f'=K{r}-M{r}', fmt='0.00000000')
    cell(ws4, r, 15, f'=IF({BTC_HOLDINGS}*{BTC_PRICE}>0,D{r}/({BTC_HOLDINGS}*{BTC_PRICE}),0)', fmt='0.0%')
    cell(ws4, r, 16, f"='Leveraged Position'!AQ{r}", fmt='0.00000000')
    cell(ws4, r, 17, f'=IF({MODE}="Accumulation","N/A",IF(D{r}<=0,"NO INCOME (BTC fell)",IF(F{r},"INCOME CAPPED",IF(D{r}<=B{r},"SUSTAINABLE","DRAINING CAPITAL"))))')

print("Tab 4 (Income/Accumulation) created.")

# ============================================================
# TAB 5: SUMMARY
# ============================================================
ws5 = wb.create_sheet("Summary")
ws5.sheet_properties.tabColor = "2E8B57"
ws5.column_dimensions['A'].width = 38
for c in range(2, 5):
    ws5.column_dimensions[get_column_letter(c)].width = 22

ws5.merge_cells('A1:D1')
cell(ws5, 1, 1, f'="SUMMARY — "&{MODE}&" Mode (Median Scenario)"', font=title_font, align='left')

# Projection period header
cell(ws5, 2, 1, 'Projection Period', font=sec_font, fill=sec_fill, align='left')
cell(ws5, 2, 2, '', font=sec_font, fill=sec_fill)
cell(ws5, 2, 3, '', font=sec_font, fill=sec_fill)
cell(ws5, 2, 4, '', font=sec_font, fill=sec_fill)

cell(ws5, 3, 1, 'Start Year', font=lbl_font, fill=input_bg, align='left')
cell(ws5, 3, 2, f'={START_YEAR}', fmt='0', font=val_font)
cell(ws5, 4, 1, 'End Year', font=lbl_font, fill=input_bg, align='left')
cell(ws5, 4, 2, f'={START_YEAR}+{PROJ_LENGTH}', fmt='0', font=val_font)
cell(ws5, 5, 1, 'Duration (years)', font=lbl_font, fill=input_bg, align='left')
cell(ws5, 5, 2, f'={PROJ_LENGTH}', fmt='0', font=val_font)

# Metrics start at row 7
METRICS_START_ROW = 7

# Define the last-row reference for dynamic formulas
LAST_ROW_REF = f'{PROJ_LENGTH}+2'

# Summary formulas intentionally branch directly on MODE for end-of-period values.
# This makes mode-sensitive summary metrics recalculate immediately when the
# Inputs tab mode toggle changes, instead of depending only on transitive
# recalculation through the year-by-year sheets.
SUMMARY_NET_BTC_END = (
    f'=IF({MODE}="Accumulation",'
    f'INDEX(\'Leveraged Position\'!U:U,{LAST_ROW_REF})/INDEX(\'Leveraged Position\'!B:B,{LAST_ROW_REF}),'
    f'INDEX(\'Leveraged Position\'!AC:AC,{LAST_ROW_REF}))'
)
SUMMARY_NET_WORTH_END = (
    f'=IF({MODE}="Accumulation",'
    f'INDEX(\'Leveraged Position\'!U:U,{LAST_ROW_REF}),'
    f'INDEX(\'Leveraged Position\'!AB:AB,{LAST_ROW_REF}))'
)

metrics = [
    ('Starting BTC', f'={BTC_HOLDINGS}', '0.00000000'),
    ('Starting Net Worth (USD)', f'={BTC_HOLDINGS}*{BTC_PRICE}', '$#,##0'),
    ('Survived Full Term?',
     f'=IF(COUNTIF(\'Leveraged Position\'!AL2:INDEX(\'Leveraged Position\'!AL:AL,{LAST_ROW_REF}),"LIQUIDATED")>0,"FAILED","SURVIVED")', None),
    ('Liquidation Year (if any)',
     f'=IFERROR(MATCH("LIQUIDATED",\'Leveraged Position\'!AL2:INDEX(\'Leveraged Position\'!AL:AL,{LAST_ROW_REF}),0)-1,"N/A")', None),
    ('Total Income Withdrawn',
     f'=IF({MODE}="Accumulation","N/A",INDEX(\'Income-Accumulation\'!G:G,{LAST_ROW_REF}))', '$#,##0'),
    ('Average Annual Income',
     f'=IF({MODE}="Accumulation","N/A",INDEX(\'Income-Accumulation\'!G:G,{LAST_ROW_REF})/{PROJ_LENGTH})', '$#,##0'),
    ('Net BTC at End (after debt)',
     SUMMARY_NET_BTC_END, '0.00000000'),
    ('vs Passive Hold (BTC)', f'={BTC_HOLDINGS}', '0.00000000'),
    ('BTC Accumulation Multiple',
     f'=({SUMMARY_NET_BTC_END[1:]})/{BTC_HOLDINGS}', '0.00x'),
    ('Net Worth at End (USD)',
     SUMMARY_NET_WORTH_END, '$#,##0'),
    ('vs Passive Hold (USD)',
     f'={BTC_HOLDINGS}*INDEX(\'Price Projection\'!H:H,{LAST_ROW_REF})', '$#,##0'),
    ('Worst Effective LTV',
     f'=MAX(\'Leveraged Position\'!AH2:INDEX(\'Leveraged Position\'!AH:AH,{LAST_ROW_REF}))', '0.0%'),
    ('Years in Warning/Red Zone',
     f'=COUNTIF(\'Leveraged Position\'!AL2:INDEX(\'Leveraged Position\'!AL:AL,{LAST_ROW_REF}),"WARNING")+COUNTIF(\'Leveraged Position\'!AL2:INDEX(\'Leveraged Position\'!AL:AL,{LAST_ROW_REF}),"MARGIN CALL")+COUNTIF(\'Leveraged Position\'!AL2:INDEX(\'Leveraged Position\'!AL:AL,{LAST_ROW_REF}),"LIQUIDATED")', None),
    ('Maximum Debt Carried',
     f'=MAX(\'Leveraged Position\'!P2:INDEX(\'Leveraged Position\'!P:P,{LAST_ROW_REF}))', '$#,##0'),
    ('Total Interest Paid',
     f'=SUM(\'Leveraged Position\'!H2:INDEX(\'Leveraged Position\'!H:H,{LAST_ROW_REF}))', '$#,##0'),
    ('Net BTC Accumulated',
     f'=({SUMMARY_NET_BTC_END[1:]})-{BTC_HOLDINGS}', '0.00000000'),
]

for i, (name, formula, fmt) in enumerate(metrics):
    r = i + METRICS_START_ROW
    cell(ws5, r, 1, name, font=lbl_font, fill=input_bg, align='left')
    cell(ws5, r, 2, formula, fmt=fmt, font=val_font)

# ============================================================
# CONDITIONAL FORMATTING & CHARTS
# ============================================================
# Tab 3: Risk status colours — use wide range to accommodate longer projections
# Data starts row 2; cover rows 2-52 (up to 50 years)
ws3.conditional_formatting.add('AL2:AL52',
    CellIsRule(operator='equal', formula=['"SAFE"'], fill=green_bg))
ws3.conditional_formatting.add('AL2:AL52',
    CellIsRule(operator='equal', formula=['"WARNING"'], fill=amber_bg))
ws3.conditional_formatting.add('AL2:AL52',
    CellIsRule(operator='equal', formula=['"MARGIN CALL"'], fill=amber_bg))
ws3.conditional_formatting.add('AL2:AL52',
    CellIsRule(operator='equal', formula=['"LIQUIDATED"'], fill=red_bg))

# Tab 4: Sustainability colours
ws4.conditional_formatting.add('Q2:Q52',
    CellIsRule(operator='equal', formula=['"DRAINING CAPITAL"'], fill=red_bg))
ws4.conditional_formatting.add('Q2:Q52',
    CellIsRule(operator='equal', formula=['"SUSTAINABLE"'], fill=green_bg))

# Chart: Price Projection (Tab 2)
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties

chart1 = LineChart()
chart1.title = "BTC Price Projection (Median Scenario)"
chart1.style = 10
chart1.height = 15
chart1.width = 35

# Data ranges match the generated projection rows
data1 = Reference(ws2, min_col=6, min_row=1, max_row=CHART_ROWS)
cats1 = Reference(ws2, min_col=2, min_row=2, max_row=CHART_ROWS)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.series[0].graphicalProperties.line.width = 25000

data2 = Reference(ws2, min_col=8, min_row=1, max_row=CHART_ROWS)
chart1.add_data(data2, titles_from_data=True)
chart1.series[1].graphicalProperties.line.width = 25000

# Axis labels — set AFTER adding data (openpyxl quirk)
chart1.y_axis.title = "USD"
chart1.y_axis.numFmt = '#,##0'
chart1.y_axis.delete = False
chart1.y_axis.axPos = "l"
chart1.x_axis.title = "Year"
chart1.x_axis.delete = False
chart1.x_axis.axPos = "b"
# Rotate x-axis labels 45 degrees to prevent overlap
chart1.x_axis.txPr = RichText(
    bodyPr=RichTextProperties(rot="-2700000", vert="horz"),
    p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=800)))]
)

ws2.add_chart(chart1, "A24")

# Chart: Net BTC Over Time (Tab 4)
chart2 = LineChart()
chart2.title = "Net BTC: Leveraged vs Passive Hold"
chart2.style = 10
chart2.height = 15
chart2.width = 35

data_a = Reference(ws4, min_col=11, min_row=1, max_row=CHART_ROWS)
data_b = Reference(ws4, min_col=13, min_row=1, max_row=CHART_ROWS)
cats2 = Reference(ws4, min_col=1, min_row=2, max_row=CHART_ROWS)
chart2.add_data(data_a, titles_from_data=True)
chart2.add_data(data_b, titles_from_data=True)
chart2.set_categories(cats2)
chart2.series[0].graphicalProperties.line.width = 25000
chart2.series[1].graphicalProperties.line.width = 25000

chart2.y_axis.title = "BTC"
chart2.y_axis.numFmt = '0.0000'
chart2.y_axis.delete = False
chart2.y_axis.axPos = "l"
chart2.x_axis.title = "Year"
chart2.x_axis.delete = False
chart2.x_axis.axPos = "b"
chart2.x_axis.txPr = RichText(
    bodyPr=RichTextProperties(rot="-2700000", vert="horz"),
    p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=800)))]
)

ws4.add_chart(chart2, "A24")

# Save
output_path = r'C:\hermes\btc_leveraged_model.xlsx'
# If file is locked (open in Excel), save with a new name
try:
    wb.save(output_path)
    print(f"\nSpreadsheet saved to: {output_path}")
except PermissionError:
    import os, glob
    existing = glob.glob(r'C:\hermes\btc_leveraged_model_v*.xlsx')
    v = len(existing) + 2
    output_path = rf'C:\hermes\btc_leveraged_model_v{v}.xlsx'
    wb.save(output_path)
    print(f"\nSpreadsheet saved to: {output_path} (original was locked)")
print("Tabs: Inputs | Price Projection | Leveraged Position | Income-Accumulation | Summary")
