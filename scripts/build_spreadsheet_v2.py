#!/usr/bin/env python3
"""Build the clean-sheet BTC-backed loan model v2 workbook."""

from __future__ import annotations

import argparse
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from model_v2.income import income_year
from model_v2.risk import effective_ltv, price_drop_to_threshold
from model_v2.scenarios import V2ProjectionResult, default_v2_inputs, run_v2_projection
from model_v2.types import BorrowTerms, IncomeConfig, IncomeRow, PricePoint, RiskThresholds

HEADER_FILL = PatternFill(start_color="023D4B", end_color="023D4B", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
INPUT_FILL = PatternFill(start_color="F6F6F6", end_color="F6F6F6", fill_type="solid")
THIN = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="023D4B")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="023D4B")
LABEL_FONT = Font(name="Calibri", size=11, color="333333")
VALUE_FONT = Font(name="Calibri", size=11, bold=True)

USD_FMT = '$#,##0.00'
BTC_FMT = '0.00000000'
PCT_FMT = '0.00%'


def _cell(ws, row: int, col: int, value=None, *, fmt: str | None = None, font=None, fill=None, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fmt:
        cell.number_format = fmt
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    return cell


def _title(ws, text: str) -> None:
    ws.merge_cells("A1:H1")
    _cell(ws, 1, 1, text, font=TITLE_FONT, align="left")


def _headers(ws, row: int, headers: list[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        _cell(ws, row, idx, header, font=HEADER_FONT, fill=HEADER_FILL)


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 28)


def _write_inputs(wb: Workbook, projection: V2ProjectionResult) -> None:
    ws = wb.active
    ws.title = "Inputs"
    _title(ws, "BTC-Backed Loan Model v2 Inputs")
    cfg = default_v2_inputs().accumulation_config
    income_cfg = default_v2_inputs().income_config

    rows = [
        ("Shared", None, None, None),
        ("Current BTC Price", cfg.current_btc_price, "USD", "Year 0 price; feeds Price Projection"),
        ("Annual BTC Price Growth", 0.0, "%", "Editable flat/default projection growth rate"),
        ("Starting BTC", cfg.starting_btc, "BTC", "Collateral used by both engines"),
        ("Borrow APR", cfg.borrow_terms.borrow_apr, "%", "Interest accrues on start-of-row debt"),
        ("Interest Treatment", str(cfg.borrow_terms.interest_treatment), "choice", "capitalized or paid externally"),
        ("Margin Call LTV", cfg.risk_thresholds.margin_call_ltv, "%", "Risk threshold"),
        ("Liquidation LTV", cfg.risk_thresholds.liquidation_ltv, "%", "Risk threshold"),
        ("Minimum Liquidation Buffer", cfg.risk_thresholds.minimum_liquidation_buffer, "%", "Required price-drop buffer"),
        ("Accumulation", None, None, None),
        ("Target Effective LTV", cfg.target_effective_ltv, "%", "Recursive setup/top-up target"),
        ("Max Effective LTV", cfg.max_effective_ltv, "%", "Hard cap"),
        ("Setup Collateral BTC", "=$B$7/(1-MIN($B$14,$B$15,$B$11*(1-$B$12)))", "BTC", "Formula output from starting BTC and effective LTV cap"),
        ("Income", None, None, None),
        ("Max Available Annual Income (Year 1)", "=MAX(0,$B$7*'Price Projection'!B5*MIN($B$20,$B$11*(1-$B$12)))", "USD", "Formula output from BTC collateral, LTV ceiling, and liquidation buffer"),
        ("Selected Annual Income Draw", income_cfg.selected_annual_income_draw, "USD", "User-selected draw; capped by max available annual income"),
        ("Income LTV Ceiling", income_cfg.income_ltv_ceiling, "%", "Safe debt ceiling"),
        ("Setup Collateral BTC", "=$B$7", "BTC", "Income keeps BTC constant"),
        ("Advanced Risk", None, None, None),
        ("Warning LTV", cfg.risk_thresholds.warning_ltv, "%", "Status changes to WARNING at or above this LTV"),
    ]
    _headers(ws, 3, ["Input", "Value", "Unit", "Notes"])
    for idx, row in enumerate(rows, start=4):
        name, value, unit, notes = row
        if value is None and unit is None:
            ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=4)
            _cell(ws, idx, 1, name, font=SECTION_FONT, fill=SECTION_FILL, align="left")
            continue
        fmt = PCT_FMT if unit == "%" else USD_FMT if unit == "USD" else BTC_FMT if unit == "BTC" else None
        _cell(ws, idx, 1, name, font=LABEL_FONT, fill=INPUT_FILL, align="left")
        _cell(ws, idx, 2, value, fmt=fmt, font=VALUE_FONT)
        _cell(ws, idx, 3, unit)
        _cell(ws, idx, 4, notes, align="left")
    _autosize(ws)


def _write_price_projection(wb: Workbook, projection: V2ProjectionResult) -> None:
    ws = wb.create_sheet("Price Projection")
    _title(ws, "Price Projection")
    _headers(ws, 3, ["Year", "BTC Price", "Scenario", "YoY Change"])
    for row_idx, point in enumerate(projection.price_points, start=4):
        _cell(ws, row_idx, 1, point.year)
        price_formula = "='Inputs'!$B$5" if point.year == 0 else f"=B{row_idx - 1}*(1+'Inputs'!$B$6)"
        yoy_formula = 0 if point.year == 0 else f"=B{row_idx}/B{row_idx - 1}-1"
        _cell(ws, row_idx, 2, price_formula, fmt=USD_FMT)
        _cell(ws, row_idx, 3, "Editable formula projection")
        _cell(ws, row_idx, 4, yoy_formula, fmt=PCT_FMT)
    _autosize(ws)


def _write_accumulation(wb: Workbook, projection: V2ProjectionResult) -> None:
    ws = wb.create_sheet("Accumulation Engine")
    _title(ws, "Accumulation Engine — Recursive BTC Leverage")
    headers = [
        "Year", "BTC Price", "Starting Collateral BTC", "Starting Debt",
        "Interest Accrued", "Debt After Interest", "Pre-Action LTV", "New Borrowing",
        "BTC Bought", "Ending Collateral BTC", "Ending Debt", "Gross BTC Held",
        "Debt BTC Equivalent", "Net BTC After Debt", "Extra BTC vs Passive",
        "Net Equity USD", "Effective LTV End", "Leverage Multiple",
        "Margin-Call Price", "Liquidation Price", "Drop to Liquidation",
        "Risk Status",
    ]
    _headers(ws, 3, headers)
    borrow_ltv = "MIN('Inputs'!$B$14,'Inputs'!$B$15,'Inputs'!$B$11*(1-'Inputs'!$B$12))"
    for row_idx, row in enumerate(projection.accumulation_rows, start=4):
        if row_idx == 4:
            starting_collateral = "='Inputs'!$B$7"
            starting_debt = 0
            interest = 0
            debt_after_interest = "=D4+E4"
            new_borrowing = f"=MAX(0,({borrow_ltv}*C4*B4-F4)/(1-{borrow_ltv}))"
        else:
            prev = row_idx - 1
            starting_collateral = f"=J{prev}"
            starting_debt = f"=K{prev}"
            interest = f"=D{row_idx}*'Inputs'!$B$8"
            debt_after_interest = f'=IF(\'Inputs\'!$B$9="capitalized",D{row_idx}+E{row_idx},D{row_idx})'
            new_borrowing = (
                f"=IF(G{row_idx}>='Inputs'!$B$10,0,"
                f"MAX(0,({borrow_ltv}*C{row_idx}*B{row_idx}-F{row_idx})/(1-{borrow_ltv})))"
            )
        values = [
            row.year, f"='Price Projection'!B{row_idx}", starting_collateral, starting_debt,
            interest, debt_after_interest, f"=IF(C{row_idx}*B{row_idx}=0,0,F{row_idx}/(C{row_idx}*B{row_idx}))", new_borrowing,
            f"=H{row_idx}/B{row_idx}", f"=C{row_idx}+I{row_idx}", f"=F{row_idx}+H{row_idx}", f"=J{row_idx}",
            f"=K{row_idx}/B{row_idx}", f"=L{row_idx}-M{row_idx}",
            f"=L{row_idx}-'Inputs'!$B$7", f"=L{row_idx}*B{row_idx}-K{row_idx}", f"=IF(J{row_idx}*B{row_idx}=0,0,K{row_idx}/(J{row_idx}*B{row_idx}))",
            f"=L{row_idx}/'Inputs'!$B$7", f"=IF(K{row_idx}=0,0,K{row_idx}/(J{row_idx}*'Inputs'!$B$10))", f"=IF(K{row_idx}=0,0,K{row_idx}/(J{row_idx}*'Inputs'!$B$11))",
            f"=IF(T{row_idx}<=0,0,1-T{row_idx}/B{row_idx})",
            f'=IF(Q{row_idx}=0,"SAFE",IF(Q{row_idx}>=\'Inputs\'!$B$11,"LIQUIDATED",IF(Q{row_idx}>=\'Inputs\'!$B$10,"MARGIN CALL",IF(OR(Q{row_idx}>=\'Inputs\'!$B$23,U{row_idx}<\'Inputs\'!$B$12),"WARNING","SAFE"))))',
        ]
        formats = [
            None, USD_FMT, BTC_FMT, USD_FMT,
            USD_FMT, USD_FMT, PCT_FMT, USD_FMT,
            BTC_FMT, BTC_FMT, USD_FMT, BTC_FMT,
            BTC_FMT, BTC_FMT, BTC_FMT, USD_FMT,
            PCT_FMT, None, USD_FMT, USD_FMT,
            PCT_FMT, None,
        ]
        for col, (value, fmt) in enumerate(zip(values, formats, strict=True), start=1):
            _cell(ws, row_idx, col, value, fmt=fmt)
    _autosize(ws)


def _write_income(wb: Workbook, projection: V2ProjectionResult) -> None:
    ws = wb.create_sheet("Income Engine")
    _title(ws, "Income Engine — Borrow-Funded Income")
    headers = [
        "Year", "BTC Price", "Starting Collateral BTC", "Starting Debt",
        "Interest Accrued", "Debt After Interest", "Selected Income Draw",
        "Max Available Annual Income", "Income Borrowed", "Unfunded Income",
        "Cumulative Income Borrowed", "Ending Debt", "Debt BTC Equivalent",
        "Net BTC After Debt", "Net Equity USD", "Effective LTV End",
        "Margin-Call Price", "Liquidation Price", "Drop to Liquidation",
        "Sustainability Status",
    ]
    _headers(ws, 3, headers)
    for row_idx, row in enumerate(projection.income_rows, start=4):
        if row_idx == 4:
            starting_debt = 0
            interest = 0
            debt_after_interest = "=D4+E4"
            selected_draw = 0
            cumulative_income = "=I4"
        else:
            starting_debt = f"=L{row_idx - 1}"
            interest = f"=D{row_idx}*'Inputs'!$B$8"
            debt_after_interest = f'=IF(\'Inputs\'!$B$9="capitalized",D{row_idx}+E{row_idx},D{row_idx})'
            selected_draw = "='Inputs'!$B$19"
            cumulative_income = f"=K{row_idx - 1}+I{row_idx}"
        available_capacity = f"=MAX(0,MIN(C{row_idx}*B{row_idx}*'Inputs'!$B$20,C{row_idx}*B{row_idx}*'Inputs'!$B$11*(1-'Inputs'!$B$12))-F{row_idx})"
        income_borrowed = f"=IF(IF(C{row_idx}*B{row_idx}=0,0,F{row_idx}/(C{row_idx}*B{row_idx}))>='Inputs'!$B$10,0,MIN(G{row_idx},H{row_idx}))"
        values = [
            row.year, f"='Price Projection'!B{row_idx}", "='Inputs'!$B$7", starting_debt,
            interest, debt_after_interest, selected_draw, available_capacity,
            income_borrowed, f"=G{row_idx}-I{row_idx}", cumulative_income,
            f"=F{row_idx}+I{row_idx}", f"=L{row_idx}/B{row_idx}", f"=C{row_idx}-M{row_idx}",
            f"=C{row_idx}*B{row_idx}-L{row_idx}", f"=IF(C{row_idx}*B{row_idx}=0,0,L{row_idx}/(C{row_idx}*B{row_idx}))", f"=IF(L{row_idx}=0,0,L{row_idx}/(C{row_idx}*'Inputs'!$B$10))",
            f"=IF(L{row_idx}=0,0,L{row_idx}/(C{row_idx}*'Inputs'!$B$11))", f"=IF(R{row_idx}<=0,0,1-R{row_idx}/B{row_idx})",
            f'=IF(P{row_idx}>=\'Inputs\'!$B$11,"LIQUIDATED",IF(P{row_idx}>=\'Inputs\'!$B$10,"MARGIN CALL",IF(OR(P{row_idx}>=\'Inputs\'!$B$23,S{row_idx}<\'Inputs\'!$B$12),"WARNING",IF(AND(G{row_idx}>0,I{row_idx}=0),"FAILED",IF(I{row_idx}<G{row_idx},"CONSTRAINED","SAFE")))))',
        ]
        formats = [
            None, USD_FMT, BTC_FMT, USD_FMT,
            USD_FMT, USD_FMT, USD_FMT, USD_FMT,
            USD_FMT, USD_FMT, USD_FMT,
            USD_FMT, BTC_FMT, BTC_FMT,
            USD_FMT, PCT_FMT, USD_FMT,
            USD_FMT, PCT_FMT, None,
        ]
        for col, (value, fmt) in enumerate(zip(values, formats, strict=True), start=1):
            _cell(ws, row_idx, col, value, fmt=fmt)
    _autosize(ws)


def _write_risk_alerts(wb: Workbook, projection: V2ProjectionResult) -> None:
    ws = wb.create_sheet("Risk Alerts")
    _title(ws, "Risk Alerts")
    metrics = [
        ("Accumulation risk path", None, None),
        ("Accumulation max effective LTV", "=MAX('Accumulation Engine'!Q4:Q14)", PCT_FMT),
        ("Accumulation min drop-to-liquidation", "=MIN('Accumulation Engine'!U4:U14)", PCT_FMT),
        ("Accumulation first non-safe year", '=IF(COUNTIF(\'Accumulation Engine\'!$V$4:$V$14,"<>SAFE")=0,"None",INDEX(\'Accumulation Engine\'!$A$4:$A$14,AGGREGATE(15,6,(ROW(\'Accumulation Engine\'!$V$4:$V$14)-ROW(\'Accumulation Engine\'!$V$4)+1)/(\'Accumulation Engine\'!$V$4:$V$14<>"SAFE"),1)))', None),
        ("Accumulation liquidation year", '=IFERROR(INDEX(\'Accumulation Engine\'!$A$4:$A$14,MATCH("LIQUIDATED",\'Accumulation Engine\'!$V$4:$V$14,0)),"None")', None),
        ("Accumulation total interest accrued", "=SUM('Accumulation Engine'!E4:E14)", USD_FMT),
        ("Accumulation total new borrowing", "=SUM('Accumulation Engine'!H4:H14)", USD_FMT),
        ("Accumulation ending net BTC", "='Accumulation Engine'!N14", BTC_FMT),
        ("Accumulation BTC outperformance vs passive", "='Accumulation Engine'!N14-'Inputs'!$B$7", BTC_FMT),
        ("Income risk path", None, None),
        ("Income max effective LTV", "=MAX('Income Engine'!P4:P14)", PCT_FMT),
        ("Income min drop-to-liquidation", "=MIN('Income Engine'!S4:S14)", PCT_FMT),
        ("Income first constrained year", '=IFERROR(INDEX(\'Income Engine\'!$A$4:$A$14,MATCH("CONSTRAINED",\'Income Engine\'!$T$4:$T$14,0)),"None")', None),
        ("Income liquidation year", '=IFERROR(INDEX(\'Income Engine\'!$A$4:$A$14,MATCH("LIQUIDATED",\'Income Engine\'!$T$4:$T$14,0)),"None")', None),
        ("Income total interest accrued", "=SUM('Income Engine'!E4:E14)", USD_FMT),
        ("Income total funded", "=SUM('Income Engine'!I4:I14)", USD_FMT),
        ("Income total shortfall", "=SUM('Income Engine'!J4:J14)", USD_FMT),
        ("Passive hold comparison", None, None),
        ("Passive BTC held", "='Inputs'!$B$7", BTC_FMT),
        ("Passive ending value", "='Inputs'!$B$7*'Price Projection'!B14", USD_FMT),
    ]
    _headers(ws, 3, ["Metric", "Value"])
    for idx, (name, value, fmt) in enumerate(metrics, start=4):
        if value is None:
            ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=2)
            _cell(ws, idx, 1, name, font=SECTION_FONT, fill=SECTION_FILL, align="left")
            continue
        _cell(ws, idx, 1, name, align="left")
        _cell(ws, idx, 2, value, fmt=fmt)
    _autosize(ws)


def _write_summary(wb: Workbook, projection: V2ProjectionResult) -> None:
    ws = wb.create_sheet("Summary")
    _title(ws, "BTC-Backed Loan Model v2 Summary")
    summary = [
        ("Passive starting BTC", "='Inputs'!$B$7", BTC_FMT),
        ("Passive ending BTC", "='Inputs'!$B$7", BTC_FMT),
        ("Passive ending value", "='Inputs'!$B$7*'Price Projection'!B14", USD_FMT),
        ("Accumulation setup gross BTC", "='Accumulation Engine'!J4", BTC_FMT),
        ("Accumulation ending gross BTC", "='Accumulation Engine'!J14", BTC_FMT),
        ("Accumulation ending debt", "='Accumulation Engine'!K14", USD_FMT),
        ("Accumulation ending net BTC", "='Accumulation Engine'!N14", BTC_FMT),
        ("Accumulation ending LTV", "='Accumulation Engine'!Q14", PCT_FMT),
        ("Accumulation status", "='Accumulation Engine'!V14", None),
        ("Income selected draw year 1", "='Inputs'!$B$19", USD_FMT),
        ("Income funded year 1", "='Income Engine'!I5", USD_FMT),
        ("Income shortfall year 1", "='Income Engine'!J5", USD_FMT),
        ("Income ending debt", "='Income Engine'!L14", USD_FMT),
        ("Income ending net BTC", "='Income Engine'!N14", BTC_FMT),
        ("Income status", "='Income Engine'!T14", None),
    ]
    _headers(ws, 3, ["Metric", "Value"])
    for idx, (name, value, fmt) in enumerate(summary, start=4):
        _cell(ws, idx, 1, name, align="left")
        _cell(ws, idx, 2, value, fmt=fmt)
    _autosize(ws)


def _write_audit_examples(wb: Workbook, projection: V2ProjectionResult) -> None:
    ws = wb.create_sheet("Audit Examples")
    _title(ws, "Audit Examples")
    acc = projection.accumulation_rows[0]

    risk = RiskThresholds(
        margin_call_ltv=0.70,
        liquidation_ltv=0.75,
        warning_ltv=0.60,
        minimum_liquidation_buffer=0.50,
    )
    income_capacity_cfg = IncomeConfig(
        starting_btc=2.0,
        current_btc_price=100_000.0,
        selected_annual_income_draw=50_000.0,
        income_ltv_ceiling=0.35,
        borrow_terms=BorrowTerms(borrow_apr=0.0),
        risk_thresholds=risk,
    )
    income_capacity_prior = IncomeRow(
        year=0,
        btc_price=100_000.0,
        collateral_btc=2.0,
        debt_usd=30_000.0,
        income_borrowed_usd=0.0,
        income_shortfall_usd=0.0,
        effective_ltv=0.15,
        liquidation_price=20_000.0,
    )
    income_capacity = income_year(
        income_capacity_cfg,
        income_capacity_prior,
        PricePoint(year=1, btc_price=100_000.0),
    )

    _headers(ws, 3, ["Audit Check", "Expected / Model Value"])
    rows = [
        (4, "Starting BTC", 2.0, BTC_FMT),
        (5, "Target effective LTV", 0.25, PCT_FMT),
        (6, "Final collateral BTC", acc.collateral_btc, BTC_FMT),
        (7, "BTC purchased", acc.btc_purchased, BTC_FMT),
        (8, "Debt", acc.debt_usd, USD_FMT),
        (9, "Effective LTV", acc.effective_ltv, PCT_FMT),
        (10, "Liquidation price", acc.liquidation_price, USD_FMT),
        (13, "Double-loop / 30.6% effective LTV example", None, None),
        (14, "Starting BTC", 2.0, BTC_FMT),
        (15, "BTC price", 60_000.0, USD_FMT),
        (16, "Gross BTC ≈", 2.883, BTC_FMT),
        (17, "Effective LTV ≈", 0.306, PCT_FMT),
        (18, "Debt ≈", 53_000.0, USD_FMT),
        (20, "Income capacity example", None, None),
        (21, "Collateral BTC", 2.0, BTC_FMT),
        (22, "BTC price", 100_000.0, USD_FMT),
        (23, "Existing debt", 30_000.0, USD_FMT),
        (24, "Income LTV ceiling", 0.35, PCT_FMT),
        (25, "Selected income draw", 50_000.0, USD_FMT),
        (26, "Max safe debt", 70_000.0, USD_FMT),
        (27, "Income borrowed", income_capacity.income_borrowed_usd, USD_FMT),
        (28, "Unfunded income", income_capacity.income_shortfall_usd, USD_FMT),
        (29, "Debt end", income_capacity.debt_usd, USD_FMT),
    ]
    for idx, name, value, fmt in rows:
        if value is None:
            ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=2)
            _cell(ws, idx, 1, name, font=SECTION_FONT, fill=SECTION_FILL, align="left")
            continue
        _cell(ws, idx, 1, name, align="left")
        _cell(ws, idx, 2, value, fmt=fmt)
    _autosize(ws)


def build_workbook_v2(output_path: str | Path = "btc_leveraged_model_v2.xlsx") -> Path:
    """Build and save the v2 workbook. Returns the saved path."""

    output_path = Path(output_path)
    projection = run_v2_projection(default_v2_inputs(projection_years=10))
    wb = Workbook()
    _write_inputs(wb, projection)
    _write_price_projection(wb, projection)
    _write_accumulation(wb, projection)
    _write_income(wb, projection)
    _write_risk_alerts(wb, projection)
    _write_summary(wb, projection)
    _write_audit_examples(wb, projection)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build the v2 BTC-backed loan model workbook.")
    parser.add_argument(
        "--output",
        "-o",
        default="btc_leveraged_model_v2.xlsx",
        help="Output .xlsx path (default: btc_leveraged_model_v2.xlsx)",
    )
    args = parser.parse_args(argv)

    path = build_workbook_v2(args.output)
    print(f"Wrote {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
