from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from model_v2.scenarios import default_v2_inputs, run_v2_projection  # noqa: E402
from scripts.build_spreadsheet import build_workbook as build_canonical_workbook  # noqa: E402
from scripts.build_spreadsheet_v2 import build_workbook_v2  # noqa: E402


def test_default_projection_contains_setup_plus_ten_annual_rows() -> None:
    projection = run_v2_projection(default_v2_inputs(projection_years=10))

    assert len(projection.price_points) == 11
    assert len(projection.accumulation_rows) == 11
    assert len(projection.income_rows) == 11
    assert projection.accumulation_rows[0].year == 0
    assert projection.income_rows[0].year == 0
    assert projection.accumulation_rows[-1].year == 10
    assert projection.income_rows[-1].year == 10


def test_default_projection_preserves_audit_setup_example() -> None:
    projection = run_v2_projection(default_v2_inputs(projection_years=1))
    row = projection.accumulation_rows[0]

    assert row.collateral_btc == pytest.approx(2.6666666667)
    assert row.btc_purchased == pytest.approx(0.6666666667)
    assert row.debt_usd == pytest.approx(40_000.0)
    assert row.effective_ltv == pytest.approx(0.25)
    assert row.liquidation_price == pytest.approx(20_000.0)


def test_default_income_projection_reports_capacity_shortfall() -> None:
    projection = run_v2_projection(default_v2_inputs(projection_years=1))
    row = projection.income_rows[1]

    assert row.income_borrowed_usd == pytest.approx(42_000.0)
    assert row.income_shortfall_usd == pytest.approx(8_000.0)
    assert row.warning == "CONSTRAINED"


def test_v2_workbook_builder_creates_expected_tabs_and_values(tmp_path: Path) -> None:
    output_path = tmp_path / "btc_model_v2.xlsx"

    build_workbook_v2(output_path)

    wb = load_workbook(output_path, data_only=False)
    assert wb.calculation.calcMode == "auto"
    assert wb.calculation.fullCalcOnLoad is True
    assert wb.calculation.forceFullCalc is True
    assert wb.sheetnames == [
        "Inputs",
        "Price Projection",
        "Accumulation Engine",
        "Income Engine",
        "Risk Alerts",
        "Summary",
        "Audit Examples",
    ]

    summary = wb["Summary"]
    assert summary["A1"].value == "BTC-Backed Loan Model v2 Summary"
    assert summary["B4"].value == "='Inputs'!$B$7"
    assert summary["B7"].value == "='Accumulation Engine'!J4"
    assert summary["B14"].value == "='Income Engine'!I5"
    assert summary["B15"].value == "='Income Engine'!J5"
    assert summary["A13"].value == "Income selected draw year 1"
    assert summary["A14"].value == "Income funded year 1"

    inputs = wb["Inputs"]
    input_labels = [inputs.cell(row=row, column=1).value for row in range(4, 27)]
    assert "Annual BTC Price Growth" in input_labels
    assert "Projection Years" in input_labels
    assert "Start Year" in input_labels
    assert "Selected Annual Income Draw" in input_labels
    assert "Max Available Annual Income (Year 1)" in input_labels
    assert "Annual Income Target" not in input_labels

    price_projection = wb["Price Projection"]
    assert price_projection["A4"].value == '=IF(0<=\'Inputs\'!$B$25,\'Inputs\'!$B$26+0,"")'
    assert price_projection["A14"].value == '=IF(10<=\'Inputs\'!$B$25,\'Inputs\'!$B$26+10,"")'
    assert price_projection["B4"].value == '=IF($A4="", "", \'Inputs\'!$B$5)'
    assert price_projection["B5"].value == '=IF($A5="","",B4*(1+\'Inputs\'!$B$6))'
    assert price_projection["D4"].value == '=IF($A4="","",0)'
    assert price_projection["D5"].value == '=IF($A5="","",B5/B4-1)'

    assert inputs["A5"].value == "Current BTC Price"
    assert inputs["A6"].value == "Annual BTC Price Growth"
    assert inputs["A25"].value == "Projection Years"
    assert inputs["A26"].value == "Start Year"
    assert inputs["B5"].value == pytest.approx(60_000.0)
    assert inputs["B6"].value == 0
    assert inputs["B25"].value == 10
    assert inputs["B26"].value == 2026

    accumulation = wb["Accumulation Engine"]
    assert accumulation["A3"].value == "Year"
    assert accumulation["A4"].value == "='Price Projection'!A4"
    assert accumulation["B4"].value == '=IF($A4="","",\'Price Projection\'!B4)'
    assert accumulation["B5"].value == '=IF($A5="","",\'Price Projection\'!B5)'
    assert accumulation["C3"].value == "Starting Collateral BTC"
    assert accumulation["D3"].value == "Starting Debt"
    assert accumulation["E3"].value == "Interest Accrued"
    assert accumulation["F3"].value == "Debt After Interest"
    assert accumulation["G3"].value == "Pre-Action LTV"
    assert accumulation["L3"].value == "Gross BTC Held"
    assert accumulation["M3"].value == "Debt BTC Equivalent"
    assert accumulation["N3"].value == "Net BTC After Debt"
    assert accumulation["O3"].value == "Extra BTC vs Passive"
    assert accumulation["P3"].value == "Net Equity USD"
    assert accumulation["Q3"].value == "Effective LTV End"
    assert accumulation["R3"].value == "Leverage Multiple"
    assert accumulation["S3"].value == "Margin-Call Price"
    assert accumulation["U3"].value == "Drop to Liquidation"
    assert accumulation["V3"].value == "Risk Status"
    assert accumulation["C4"].value == '=IF($A4="","",\'Inputs\'!$B$7)'
    assert accumulation["J4"].value == '=IF($A4="","",C4+I4)'
    assert accumulation["D5"].value == '=IF($A5="","",K4)'
    assert accumulation["E5"].value == '=IF($A5="","",D5*\'Inputs\'!$B$8)'
    assert "capitalized" in accumulation["F5"].value
    assert accumulation["G5"].value == '=IF($A5="","",IF(C5*B5=0,0,F5/(C5*B5)))'
    assert accumulation["Q5"].value == '=IF($A5="","",IF(J5*B5=0,0,K5/(J5*B5)))'
    assert accumulation["R5"].value == '=IF($A5="","",L5/\'Inputs\'!$B$7)'

    income = wb["Income Engine"]
    assert income["D3"].value == "Starting Debt"
    assert income["E3"].value == "Interest Accrued"
    assert income["F3"].value == "Debt After Interest"
    assert income["G3"].value == "Selected Income Draw"
    assert income["H3"].value == "Max Available Annual Income"
    assert income["J3"].value == "Unfunded Income"
    assert income["K3"].value == "Cumulative Income Borrowed"
    assert income["N3"].value == "Net BTC After Debt"
    assert income["O3"].value == "Net Equity USD"
    assert income["Q3"].value == "Margin-Call Price"
    assert income["S3"].value == "Drop to Liquidation"
    assert income["T3"].value == "Sustainability Status"
    assert income["A5"].value == "='Price Projection'!A5"
    assert income["B5"].value == '=IF($A5="","",\'Price Projection\'!B5)'
    assert income["D5"].value == '=IF($A5="","",L4)'
    assert income["G5"].value == '=IF($A5="","",\'Inputs\'!$B$19)'
    assert income["H5"].value == '=IF($A5="","",MAX(0,MIN(C5*B5*\'Inputs\'!$B$20,C5*B5*\'Inputs\'!$B$11*(1-\'Inputs\'!$B$12))-F5))'
    assert income["K5"].value == '=IF($A5="","",K4+I5)'
    assert income["T5"].value.startswith("=IF(")

    risk = wb["Risk Alerts"]
    risk_metrics = [risk.cell(row=row, column=1).value for row in range(4, 20)]
    assert "Accumulation min drop-to-liquidation" in risk_metrics
    assert "Accumulation total new borrowing" in risk_metrics
    assert "Accumulation BTC outperformance vs passive" in risk_metrics
    assert "Income min drop-to-liquidation" in risk_metrics
    assert "Income liquidation year" in risk_metrics
    assert "Income total interest accrued" in risk_metrics
    assert risk["B5"].value == "=MAX('Accumulation Engine'!Q4:Q34)"
    assert risk["B9"].value == "=SUM('Accumulation Engine'!E4:E34)"
    assert risk["B14"].value == "=MAX('Income Engine'!P4:P34)"

    assert "Price Projection" in summary["B6"].value and "LOOKUP" in summary["B6"].value
    assert summary["B13"].value == "='Inputs'!$B$19"

    audit = wb["Audit Examples"]
    assert audit["B6"].value == pytest.approx(2.6666666667)
    assert audit["B8"].value == pytest.approx(40_000.0)
    assert audit["B10"].value == pytest.approx(20_000.0)
    assert audit["A13"].value == "Double-loop / 30.6% effective LTV example"
    assert audit["B17"].value == pytest.approx(0.306)
    assert audit["A20"].value == "Income capacity example"
    assert audit["A25"].value == "Selected income draw"
    assert audit["B27"].value == pytest.approx(40_000.0)
    assert audit["B28"].value == pytest.approx(10_000.0)


def test_canonical_build_spreadsheet_script_builds_v2_without_legacy_income_label(tmp_path: Path) -> None:
    output_path = tmp_path / "canonical.xlsx"

    build_canonical_workbook(output_path)

    wb = load_workbook(output_path, data_only=True)
    assert "Income Engine" in wb.sheetnames
    inputs = wb["Inputs"]
    labels = [inputs.cell(row=row, column=1).value for row in range(1, 50)]
    assert "Selected Annual Income Draw" in labels
    assert "Max Available Annual Income (Year 1)" in labels
    assert "Annual Income Target" not in labels
